# Dump the AI-cliche mega-DB to 6 distribution formats. Run with C:\Python314\python.exe
import json, os, csv, sqlite3, html

DIR = r'D:\Users\gamma\Documents\Claude\Projects\ai-cliche-db'
OUT = os.path.join(DIR, 'exports')
os.makedirs(OUT, exist_ok=True)
BASE = os.path.join(OUT, 'ai-cliche-db')

data = json.load(open(os.path.join(DIR, 'ai-cliche-megadb.json'), encoding='utf-8-sig'))
meta = data.get('meta', {})
rows = data['entries']

COLS = ['term', 'type', 'severity', 'domain', 'category', 'platforms', 'why', 'example', 'pattern', 'source']
def norm(e):
    return {
        'term': e.get('term', ''), 'type': e.get('type', ''), 'severity': e.get('severity', ''),
        'domain': e.get('domain', ''), 'category': e.get('category', ''),
        'platforms': '; '.join(e.get('platforms', []) or []),
        'why': (e.get('why', '') or '').strip(),
        'example': e.get('example', '') or '',
        'pattern': e.get('pattern', '') or '',
        'source': '; '.join(e.get('sources', []) or []) or e.get('source', ''),
    }
R = [norm(e) for e in rows]
SEV = {'high': 0, 'medium': 1, 'low': 2, '': 3}
R.sort(key=lambda r: (r['domain'], r['type'], SEV.get(r['severity'], 3), r['term'].lower()))

# ---------- 1. CSV ----------
with open(BASE + '.csv', 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.DictWriter(f, fieldnames=COLS); w.writeheader()
    for r in R: w.writerow(r)

# ---------- 2. JSON ----------
json.dump({'meta': meta, 'entries': rows}, open(BASE + '.json', 'w', encoding='utf-8'),
          ensure_ascii=False, indent=2)

# ---------- 3. TXT (grep-friendly) ----------
with open(BASE + '.txt', 'w', encoding='utf-8') as f:
    f.write(f"AI-CLICHE DETECTION DB  -  {len(R)} entries\n")
    f.write("format: [severity] domain/type | category | TERM :: why\n")
    f.write("=" * 78 + "\n\n")
    for r in R:
        f.write(f"[{r['severity'] or '-':6}] {r['domain']}/{r['type']} | {r['category']} | "
                f"{r['term']} :: {r['why']}\n")

# ---------- 4. SQLite (.db) indexed + canonical view ----------
dbp = BASE + '.db'
if os.path.exists(dbp): os.remove(dbp)
con = sqlite3.connect(dbp); cur = con.cursor()
cur.execute("""CREATE TABLE tells(
  id INTEGER PRIMARY KEY, term TEXT, type TEXT, severity TEXT, domain TEXT,
  category TEXT, platforms TEXT, why TEXT, example TEXT, pattern TEXT, source TEXT)""")
cur.executemany(
  "INSERT INTO tells(term,type,severity,domain,category,platforms,why,example,pattern,source) "
  "VALUES(:term,:type,:severity,:domain,:category,:platforms,:why,:example,:pattern,:source)", R)
for col in ('term', 'type', 'severity', 'domain', 'category', 'source'):
    cur.execute(f"CREATE INDEX idx_tells_{col} ON tells({col})")
# canonical view: the clean, matchable detection core (drops low-severity + non-matchable structures)
cur.execute("""CREATE VIEW canonical_tells AS
  SELECT term, type, severity, domain, category, why, pattern, source FROM tells
  WHERE type IN ('word','phrase','rhyme_pair','trope','opener','closer','template')
    AND severity IN ('high','medium')
  ORDER BY domain, type, CASE severity WHEN 'high' THEN 0 ELSE 1 END, term""")
cur.execute("""CREATE VIEW category_counts AS
  SELECT domain, type, COUNT(*) AS n FROM tells GROUP BY domain, type ORDER BY n DESC""")
con.commit()

# ---------- 5. XLSX ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
wb = Workbook()
ws = wb.active; ws.title = 'tells'
hdr_fill = PatternFill('solid', fgColor='1F2735'); hdr_font = Font(bold=True, color='FFFFFF')
ws.append([c.upper() for c in COLS])
for c in range(1, len(COLS) + 1):
    cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font
sevfill = {'high': PatternFill('solid', fgColor='FFD6DD'),
           'medium': PatternFill('solid', fgColor='FFEBCC'),
           'low': PatternFill('solid', fgColor='FFF6CC')}
for r in R:
    ws.append([r[c] for c in COLS])
    sc = ws.cell(ws.max_row, 3)
    if r['severity'] in sevfill: sc.fill = sevfill[r['severity']]
widths = {'term': 30, 'type': 11, 'severity': 9, 'domain': 8, 'category': 24,
          'platforms': 20, 'why': 70, 'example': 28, 'pattern': 34, 'source': 16}
for i, c in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = widths[c]
ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions
ws['G1'].alignment = Alignment(wrap_text=False)
# summary sheet
ws2 = wb.create_sheet('by_class')
ws2.append(['DOMAIN', 'TYPE', 'COUNT'])
for cell in ws2[1]: cell.font = hdr_font; cell.fill = hdr_fill
for domain, typ, n in cur.execute("SELECT domain,type,COUNT(*) FROM tells GROUP BY domain,type ORDER BY domain,type"):
    ws2.append([domain, typ, n])
ws2.column_dimensions['A'].width = 10; ws2.column_dimensions['B'].width = 12; ws2.column_dimensions['C'].width = 8
# readme sheet
ws3 = wb.create_sheet('README')
ws3.append(['AI-Cliche Detection Database'])
ws3.append([f"{len(R)} entries"])
ws3.append(['Overused AI words / phrases / rhyme-pairs / tropes / regex pattern-detectors that flag text or lyrics as machine-generated.'])
ws3.append(['Sources: 26-agent research workflow, Wikipedia Signs-of-AI-writing, EQ-Bench antislop, Detect-AI, fiction/RP slop, per-platform tells.'])
ws3.column_dimensions['A'].width = 120
wb.save(BASE + '.xlsx')

# ---------- 6. PDF (two-column, grouped by class=type) ----------
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import BaseDocTemplate, PageTemplate, Frame, Paragraph, Spacer, FrameBreak

styles = getSampleStyleSheet()
h_style = ParagraphStyle('grp', parent=styles['Heading3'], fontSize=10, spaceBefore=6, spaceAfter=3,
                         textColor=colors.HexColor('#0b5fa5'))
e_style = ParagraphStyle('ent', parent=styles['Normal'], fontSize=6.6, leading=8.2, spaceAfter=1.2)
title_style = ParagraphStyle('ttl', parent=styles['Title'], fontSize=16)

pw, ph = letter
m = 1.2 * cm; gap = 0.6 * cm
fw = (pw - 2 * m - gap) / 2
fh = ph - 2 * m - 1.0 * cm
f1 = Frame(m, m, fw, fh, leftPadding=2, rightPadding=4, topPadding=2, bottomPadding=2)
f2 = Frame(m + fw + gap, m, fw, fh, leftPadding=2, rightPadding=4, topPadding=2, bottomPadding=2)

def header(canvas, doc):
    canvas.saveState(); canvas.setFont('Helvetica', 7); canvas.setFillColor(colors.grey)
    canvas.drawString(m, ph - m + 6, f"AI-Cliche Detection DB  -  {len(R)} entries  (grouped by class)")
    canvas.drawRightString(pw - m, ph - m + 6, f"page {doc.page}")
    canvas.restoreState()

doc = BaseDocTemplate(BASE + '.pdf', pagesize=letter,
                      pageTemplates=[PageTemplate(id='2col', frames=[f1, f2], onPage=header)])
TYPE_ORDER = ['word', 'phrase', 'rhyme_pair', 'trope', 'template', 'opener', 'closer', 'structure']
def esc(s, n=120):
    s = (s or '')[:n]
    return html.escape(s)
story = []
groups = {}
for r in R: groups.setdefault(r['type'], []).append(r)
for typ in TYPE_ORDER + [t for t in groups if t not in TYPE_ORDER]:
    if typ not in groups: continue
    g = sorted(groups[typ], key=lambda r: (SEV.get(r['severity'], 3), r['term'].lower()))
    story.append(Paragraph(f"{typ.upper()}  ({len(g)})", h_style))
    for r in g:
        sevc = {'high': '#c0392b', 'medium': '#d68910', 'low': '#b7950b'}.get(r['severity'], '#888')
        story.append(Paragraph(
            f"<b>{esc(r['term'],60)}</b> <font size=5 color='{sevc}'>[{r['severity'] or '-'}]</font> "
            f"<font size=6 color='#555'>{esc(r['why'],110)}</font>", e_style))
    story.append(Spacer(1, 3))
doc.build(story)
con.close()

# ---------- verify ----------
sizes = {os.path.basename(BASE) + ext: os.path.getsize(BASE + ext)
         for ext in ('.csv', '.json', '.txt', '.db', '.xlsx', '.pdf')}
con2 = sqlite3.connect(dbp); c2 = con2.cursor()
ntab = c2.execute("SELECT COUNT(*) FROM tells").fetchone()[0]
ncanon = c2.execute("SELECT COUNT(*) FROM canonical_tells").fetchone()[0]
views = [r[0] for r in c2.execute("SELECT name FROM sqlite_master WHERE type='view'")]
con2.close()
from openpyxl import load_workbook
wbv = load_workbook(BASE + '.xlsx', read_only=True)
xrows = wbv['tells'].max_row
pdf_head = open(BASE + '.pdf', 'rb').read(8)[:5]
pdf_tail = open(BASE + '.pdf', 'rb').read()[-8:]
print("entries:", len(R))
print("file sizes:", json.dumps(sizes, indent=1))
print("sqlite: tells rows =", ntab, "| canonical_tells rows =", ncanon, "| views =", views)
print("xlsx: sheets =", wbv.sheetnames, "| tells rows (incl header) =", xrows)
print("pdf: header =", pdf_head, "| tail =", pdf_tail, "| intact =", pdf_head == b'%PDF-' and b'EOF' in pdf_tail)
